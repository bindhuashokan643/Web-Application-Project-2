from openpyxl import load_workbook

def get_login_data(file_path):

    workbook = load_workbook(file_path)


    sheet = workbook["Login Data"]

    data = []

    for row in range(2, sheet.max_row + 1):

        username = sheet.cell(row=row, column=6).value
        password = sheet.cell(row=row, column=7).value
        expected = sheet.cell(row=row, column=8).value

     #skip empty rows
        if username is None: continue

        data.append((username, password, expected))

    return data